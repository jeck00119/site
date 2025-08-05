import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Side, Border, Font, Alignment

from src.metaclasses.singleton import Singleton


class CapabilityService(metaclass=Singleton):
    def __init__(self):
        self._row_index = 1

        self._color_green = PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='D2EBD8'))
        self._color_red = PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='FEB7B3'))

    def init_excel_headers(self, inspection_list):
        self._workbook = openpyxl.Workbook()
        self._work_sheet = self._workbook['Sheet']
        self._work_sheet.title = 'Sheet Title'

        for col, key in zip(range(0, len(inspection_list.keys())), list(inspection_list.keys())):
            self._work_sheet.column_dimensions['A'].width = 20
            self._work_sheet.cell(row=1, column=col + 1).value = key
            self._work_sheet.cell(row=1, column=col + 1).alignment = Alignment(horizontal="center",
                                                                               vertical="center")
            self._work_sheet.cell(row=1, column=col + 1).font = Font(name="Arial", size=13, color="FFFFFF")
            self._work_sheet.cell(row=1, column=col + 1).fill = PatternFill(start_color="217346",
                                                                            end_color="217346",
                                                                            fill_type="solid")

    def save_excel_report(self):
        count = 0

        last_row = self._row_index + 1
        min_row = last_row + 2
        max_row = last_row + 3
        result_row = last_row + 4


        for col in self._work_sheet.columns:
            count += 1

            col_name = str(col[0])[str(col[0]).index('.') + 1:-2]

            self._work_sheet.cell(row=min_row, column=count).value = '= MIN(' + col_name + str(
                2) + ':' + col_name + str(last_row) + ')'

            self._work_sheet.cell(row=max_row, column=count).value = '= MAX(' + col_name + str(
                2) + ':' + col_name + str(last_row) + ')'

            self._work_sheet.cell(row=result_row, column=count).value = '= ABS(' + col_name + str(
                last_row + 2) + '-' + col_name + str(last_row + 3) + ')'

            self._work_sheet.cell(row=result_row, column=count).fill = self._color_green
            rule = CellIsRule(operator='greaterThan', formula=['0.05'], fill=PatternFill(start_color='EE1111', end_color='EE1111'))

            self._work_sheet.conditional_formatting.add(col_name + str(result_row), rule)

        import os
        # Save to reports directory with proper path handling
        report_path = os.path.join(os.getcwd(), "reports", "capability_report.xlsx")
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        self._workbook.save(report_path)
        self._row_index = 1

    def add_row_to_excel(self, inspection_list):

        for col, key in zip(range(0, len(inspection_list.keys())), list(inspection_list.keys())):
            self._work_sheet.cell(row=self._row_index + 1, column=col + 1).value = \
                inspection_list[key].measurement
            self._work_sheet.cell(row=self._row_index + 1, column=col + 1).alignment = \
                Alignment(horizontal="center", vertical="center")

            if inspection_list[key].result:
                self._work_sheet.cell(row=self._row_index + 1, column=col + 1).fill = self._color_red
            else:
                self._work_sheet.cell(row=self._row_index + 1, column=col + 1).fill = self._color_red

        self._row_index += 1

