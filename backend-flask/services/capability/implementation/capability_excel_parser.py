

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Side, Border, Font, Alignment


class ITACExcelParser:
    def __init__(self):
        self._row_index = 1

    def init_excel_headers(self, inspection_list):
        self.workbook_pinsReport = openpyxl.Workbook()
        self.pinsReport_sheet = self.workbook_pinsReport['Sheet']
        self.pinsReport_sheet.title = 'Sheet Title'


        for col, key in zip(range(0, len(inspection_list.keys())), list(inspection_list.keys())):
            self.pinsReport_sheet.column_dimensions['A'].width = 20
            self.pinsReport_sheet.cell(row=1, column=col + 1).value = key
            self.pinsReport_sheet.cell(row=1, column=col + 1).alignment = Alignment(horizontal="center",
                                                                                    vertical="center")
            self.pinsReport_sheet.cell(row=1, column=col + 1).font = Font(name="Arial", size=13, color="FFFFFF")
            self.pinsReport_sheet.cell(row=1, column=col + 1).fill = PatternFill(start_color="217346",
                                                                                 end_color="217346",
                                                                                 fill_type="solid")

    def save_excel_report(self):
        self._row_index = 1
        import os
        # Save to reports directory with proper path handling
        report_path = os.path.join(os.getcwd(), "reports", "capability_report.xlsx")
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        self.workbook_pinsReport.save(report_path)

    def add_row_to_excel(self, inspection_list):
        color_green = PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='D2EBD8'))
        color_red = PatternFill(patternType='solid', bgColor=openpyxl.styles.colors.Color(rgb='FEB7B3'))

        for col, key in zip(range(0, len(inspection_list.keys())), list(inspection_list.keys())):
            self.pinsReport_sheet.cell(row=self._row_index + 1, column=col + 1).value = \
                inspection_list[key].measurement
            self.pinsReport_sheet.cell(row=self._row_index + 1, column=col + 1).alignment = \
                Alignment( horizontal="center", vertical="center")

            if inspection_list[key].result:
                self.pinsReport_sheet.cell(row=self._row_index + 1, column=col + 1).fill = color_green
            else:
                self.pinsReport_sheet.cell(row=self._row_index + 1, column=col + 1).fill = color_red

        self._row_index += 1

